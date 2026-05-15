"""
Utilities to re-read exit-type values from the Latest Earnings workbook.

Used by exit scripts so that Live_Trade_Info D/E columns are refreshed from the
Latest Earnings Document just before placing orders.

Reads via openpyxl when the file is closed. If the workbook is open in Excel
(Windows file lock), falls back to xlwings without closing the user's workbook.
"""

from __future__ import annotations

import os
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


def _base_dir() -> str:
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


LATEST_EARNINGS_FILE = os.path.join(_base_dir(), "! -- Latest Earnings Document.xlsx")
LATEST_EARNINGS_SHEET = "Trades"

# Stage_Trades_Auto uses:
#   HEADER_ROW = 3 (data starts at row 4)
#   Column O = flag
#   Column A = ticker
#   Column AA = ToS exit
#   Column AB = IBKR exit
HEADER_ROW = 3
COL_FLAG = "O"
COL_TICKER = "A"
COL_TOS_EXIT = "AA"
COL_IBKR_EXIT = "AB"


def _parse_flag(cell_value: Any) -> Optional[str | int]:
    """Return 'T' for T-rows, 1..5 for weekday rows, else None."""
    if cell_value is None:
        return None
    s = str(cell_value).strip().upper()
    if s == "T":
        return "T"
    try:
        n = int(cell_value) if isinstance(cell_value, (int, float)) else int(s)
    except (ValueError, TypeError):
        return None
    if 1 <= n <= 5:
        return n
    return None


def _cell_to_clean_str(cell_value: Any) -> str:
    """Return a trimmed string; blank if None."""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def _flatten_column_range(value: Any) -> List[Any]:
    """Normalize an xlwings column range to a flat list of cell values."""
    if value is None:
        return []
    if not isinstance(value, list):
        return [value]
    if not value:
        return []
    if isinstance(value[0], list):
        return [row[0] if row else None for row in value]
    return value


def _build_exit_type_lookup(
    o_vals: List[Any],
    a_vals: List[Any],
    aa_vals: List[Any],
    ab_vals: List[Any],
    *,
    single_day: bool,
) -> Dict[str, Dict[str, str]]:
    row_count = len(o_vals)
    lookup: Dict[str, Dict[str, str]] = {}
    for i in range(row_count):
        flag_val = _parse_flag(o_vals[i])
        if flag_val is None:
            continue

        if single_day:
            if flag_val != "T":
                continue
        else:
            if not isinstance(flag_val, int):
                continue

        ticker_val = a_vals[i] if i < len(a_vals) else None
        if ticker_val is None or str(ticker_val).strip() == "":
            continue
        ticker = str(ticker_val).strip().upper()

        tos_val = aa_vals[i] if i < len(aa_vals) else None
        ib_val = ab_vals[i] if i < len(ab_vals) else None
        lookup[ticker] = {
            "tos_exit": _cell_to_clean_str(tos_val),
            "ib_exit": _cell_to_clean_str(ib_val),
        }
    return lookup


def _read_columns_openpyxl(
    earnings_file: str,
    earnings_sheet: str,
) -> Tuple[List[Any], List[Any], List[Any], List[Any]]:
    wb = load_workbook(earnings_file, data_only=True, read_only=True)
    try:
        try:
            ws = wb[earnings_sheet]
        except KeyError:
            raise KeyError(f"Worksheet '{earnings_sheet}' not found in {earnings_file}")

        start_row = HEADER_ROW + 1
        max_row = ws.max_row or (HEADER_ROW + 1)

        if max_row < start_row:
            return [], [], [], []

        o_vals = [row[0].value for row in ws[f"{COL_FLAG}{start_row}:{COL_FLAG}{max_row}"]]
        a_vals = [row[0].value for row in ws[f"{COL_TICKER}{start_row}:{COL_TICKER}{max_row}"]]
        aa_vals = [row[0].value for row in ws[f"{COL_TOS_EXIT}{start_row}:{COL_TOS_EXIT}{max_row}"]]
        ab_vals = [row[0].value for row in ws[f"{COL_IBKR_EXIT}{start_row}:{COL_IBKR_EXIT}{max_row}"]]
        return o_vals, a_vals, aa_vals, ab_vals
    finally:
        wb.close()


def _same_workbook_path(book_path: str, target_path: str) -> bool:
    if not book_path:
        return False
    return os.path.normcase(os.path.abspath(book_path)) == os.path.normcase(
        os.path.abspath(target_path)
    )


def _find_open_xlwings_book(earnings_file: str) -> Any:
    import xlwings as xw

    abs_path = os.path.abspath(earnings_file)
    for book in xw.books:
        try:
            if _same_workbook_path(book.fullname, abs_path):
                return book
        except Exception:
            continue
    try:
        return xw.Book(abs_path)
    except Exception:
        return None


def _read_columns_xlwings(
    earnings_file: str,
    earnings_sheet: str,
) -> Tuple[List[Any], List[Any], List[Any], List[Any]]:
    try:
        import xlwings as xw
    except ImportError as e:
        raise PermissionError(
            "Latest Earnings Document is open in Excel and xlwings is not installed. "
            "Install xlwings or close the workbook and try again."
        ) from e

    abs_path = os.path.abspath(earnings_file)
    wb = _find_open_xlwings_book(abs_path)
    owned_app = False
    owned_book = False
    app = None

    if wb is None:
        app = xw.App(visible=False)
        owned_app = True
        try:
            wb = app.books.open(abs_path, read_only=True)
            owned_book = True
        except Exception as e:
            try:
                app.quit()
            except Exception:
                pass
            raise PermissionError(
                "Could not read Latest Earnings Document (file may be locked). "
                "Please close it in Excel and try again."
            ) from e

    try:
        try:
            sheet = wb.sheets[earnings_sheet]
        except Exception as e:
            raise KeyError(
                f"Worksheet '{earnings_sheet}' not found in {earnings_file}"
            ) from e

        start_row = HEADER_ROW + 1
        try:
            max_row = int(sheet.used_range.last_cell.row)
        except Exception:
            max_row = start_row

        if max_row < start_row:
            return [], [], [], []

        o_vals = _flatten_column_range(
            sheet.range(f"{COL_FLAG}{start_row}:{COL_FLAG}{max_row}").value
        )
        a_vals = _flatten_column_range(
            sheet.range(f"{COL_TICKER}{start_row}:{COL_TICKER}{max_row}").value
        )
        aa_vals = _flatten_column_range(
            sheet.range(f"{COL_TOS_EXIT}{start_row}:{COL_TOS_EXIT}{max_row}").value
        )
        ab_vals = _flatten_column_range(
            sheet.range(f"{COL_IBKR_EXIT}{start_row}:{COL_IBKR_EXIT}{max_row}").value
        )
        return o_vals, a_vals, aa_vals, ab_vals
    finally:
        if owned_book and wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        if owned_app and app is not None:
            try:
                app.quit()
            except Exception:
                pass


def read_exit_types_from_latest_earnings(
    *,
    single_day: bool,
    earnings_file: str = LATEST_EARNINGS_FILE,
    earnings_sheet: str = LATEST_EARNINGS_SHEET,
) -> Dict[str, Dict[str, str]]:
    """
    Read exit-type values from Latest Earnings Document.

    Returns a dict keyed by ticker_upper:
        {
          "AAPL": {"tos_exit": "...", "ib_exit": "..."},
          ...
        }

    Filtering:
    - single_day=True  -> include only rows where flag (col O) == "T"
    - single_day=False -> include only rows where flag (col O) is 1..5

    Matching note:
    - If the same ticker appears multiple times in the filtered range, last one wins.

    Uses openpyxl when the file is not open in Excel; uses xlwings when the
    workbook is already open (Windows file lock).
    """
    if not os.path.exists(earnings_file):
        raise FileNotFoundError(f"Latest earnings file not found: {earnings_file}")

    try:
        o_vals, a_vals, aa_vals, ab_vals = _read_columns_openpyxl(
            earnings_file, earnings_sheet
        )
    except PermissionError:
        try:
            o_vals, a_vals, aa_vals, ab_vals = _read_columns_xlwings(
                earnings_file, earnings_sheet
            )
        except PermissionError:
            raise
        except KeyError:
            raise
        except Exception as e:
            raise PermissionError(
                "Could not read Latest Earnings Document while it is open in Excel. "
                "Please save and close the workbook, or ensure xlwings can connect to Excel."
            ) from e

    return _build_exit_type_lookup(
        o_vals, a_vals, aa_vals, ab_vals, single_day=single_day
    )
