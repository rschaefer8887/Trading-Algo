"""
Write Opening Prices

Reads tickers from the Latest Earnings workbook, fetches opening prices via
yfinance, writes to Opening_Prices_One_Time.xlsx and into the Earnings
workbook column T (same row as each ticker).

The Earnings file is opened and saved via xlwings (Excel) so that external
links and other Excel features are preserved. Close the workbook in Excel
before running.
"""

import os
import warnings
from datetime import datetime

import yfinance as yf
from openpyxl import Workbook, load_workbook

warnings.filterwarnings("ignore", message=".*Unknown extension.*", category=UserWarning)
warnings.filterwarnings("ignore", message=".*Conditional Formatting extension.*", category=UserWarning)

try:
    import xlwings as xw
except ImportError:
    xw = None

# ---------------------------------------------------------------------------
# Paths: Excel files live in the Trading Algo folder (parent of this script)
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR = os.path.dirname(_SCRIPT_DIR)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
source_file = os.path.join(_BASE_DIR, "! -- Latest Earnings Document.xlsx")
source_sheet = "Trades"
header_row = 3
ticker_column_letter = "A"

# Column in the Earnings workbook where opening price will be written (same row as ticker)
opening_price_column_letter = "T"

output_file = os.path.join(_BASE_DIR, "Opening_Prices_One_Time.xlsx")
output_sheet = "Prices"

# ---------------------------------------------------------------------------
# Get the row range to process
# ---------------------------------------------------------------------------
start_row = int(input(f"Enter starting row number to process (first ticker is row {header_row+1}): "))
end_row = int(input("Enter ending row number to process: "))

if end_row < start_row:
    print("Ending row must be >= starting row.")
    exit()

# ---------------------------------------------------------------------------
# Use xlwings to open Earnings in Excel, read tickers, write opening prices, save
# (Excel does the save so external links are preserved)
# ---------------------------------------------------------------------------
if xw is None:
    print("xlwings is not installed. Install it with: pip install xlwings")
    exit(1)

if not os.path.exists(source_file):
    print(f"Earnings file not found: {source_file}")
    exit(1)

app = None
try:
    app = xw.App(visible=False)
    wb_earnings = app.books.open(os.path.abspath(source_file))
    try:
        sheet = wb_earnings.sheets[source_sheet]
    except Exception:
        print(f"Worksheet '{source_sheet}' not found in {source_file}.")
        wb_earnings.close()
        app.quit()
        exit(1)

    # Build list of (row, ticker_normalized) for rows that have a ticker
    row_tickers = []
    for row in range(start_row, end_row + 1):
        cell = sheet.range(f"{ticker_column_letter}{row}").value
        if cell is None:
            continue
        ticker = str(cell).strip().upper().replace(".", "-")
        row_tickers.append((row, ticker))

    if not row_tickers:
        print("No tickers found in the specified rows.")
        wb_earnings.close()
        app.quit()
        exit(1)

    tickers = [t for _, t in row_tickers]
    print(f"Found {len(tickers)} tickers: {tickers}")

    # Fetch latest opening prices from Yahoo Finance
    data = yf.download(tickers, period="1d", group_by='ticker', auto_adjust=True, threads=True)

    if data.empty:
        print("No data returned from yfinance!")
        wb_earnings.close()
        app.quit()
        exit(1)

    # Build ticker -> opening price (rounded to 2 decimals)
    price_map = {}
    for ticker in tickers:
        try:
            if len(tickers) > 1:
                price = data[ticker]['Open'].iloc[-1]
            else:
                price = data['Open'].iloc[-1]
            price_map[ticker] = round(float(price), 2) if price is not None else None
        except Exception as e:
            print(f"Could not get price for {ticker}: {e}")
            price_map[ticker] = None

    # Write opening prices to column T (Excel saves the file → preserves external links)
    for row, ticker in row_tickers:
        val = price_map.get(ticker)
        sheet.range(f"{opening_price_column_letter}{row}").value = val

    wb_earnings.save()
    wb_earnings.close()
    print(f"Wrote opening prices to column {opening_price_column_letter} in '{source_file}' and saved (via Excel).")
finally:
    if app is not None:
        app.quit()

# ---------------------------------------------------------------------------
# Also write to Opening_Prices_One_Time.xlsx (same as Get_Opening_Prices)
# ---------------------------------------------------------------------------
if os.path.exists(output_file):
    wb_output = load_workbook(output_file)
    ws_output = wb_output.active
    ws_output.title = output_sheet
    ws_output.delete_rows(1, ws_output.max_row)
else:
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = output_sheet

ws_output.append(["Ticker", "Date", "Opening Price"])
today = datetime.today().strftime("%Y-%m-%d")

for ticker in tickers:
    price = price_map.get(ticker)
    ws_output.append([ticker, today, price])

wb_output.save(output_file)
print(f"Successfully wrote {len(tickers)} tickers to '{output_file}'")
