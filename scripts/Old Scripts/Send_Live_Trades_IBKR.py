"""
Send Live Trades to Interactive Brokers based on Live_Trade_Info.xlsx

Workflow:
- Reads tickers, direction, and share size from Live_Trade_Info.xlsx
- Builds **market-on-close (MOC)** orders (BUY for long, SELL for short)
- Connects to Interactive Brokers via TWS / IB Gateway API
- Places the orders (or just prints them when DRY_RUN is enabled)

Prerequisites:
- IBKR Desktop / TWS or IB Gateway running with API enabled.
- Python package 'ib_insync' installed:
    pip install ib_insync
"""

import os
import asyncio
import warnings
from typing import List, Tuple

from openpyxl import load_workbook

warnings.filterwarnings("ignore", message=".*Unknown extension.*", category=UserWarning)
warnings.filterwarnings("ignore", message=".*Conditional Formatting extension.*", category=UserWarning)

# Ensure there is an asyncio event loop for ib_insync/eventkit on newer Python versions.
# Some newer Python builds (like 3.11+) start without a default loop in the main thread,
# which causes eventkit/ib_insync imports to fail unless we create one first.
try:
    asyncio.get_event_loop()
except RuntimeError:
    asyncio.set_event_loop(asyncio.new_event_loop())

# Try importing ib_insync; keep the original error so we can show it
IB_IMPORT_ERROR = None
try:
    from ib_insync import IB, Stock, Order
except Exception as e:  # catch any import-time problem
    IB = None  # type: ignore[assignment]
    IB_IMPORT_ERROR = e


# ---------------------------------------------------------------------------
# Paths: Excel files live in repo root (Trading Algo); script is in Old Scripts
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR = os.path.dirname(os.path.dirname(_SCRIPT_DIR))

# ---------------------------------------------------------------------------
# Configuration — adjust to your environment
# ---------------------------------------------------------------------------
# Path to the Excel file produced by Obtain_Live_Trade_Info.py.
# This file is the single source of truth for which trades to send.
LIVE_INFO_FILE = os.path.join(_BASE_DIR, "Live_Trade_Info.xlsx")
LIVE_INFO_SHEET = "Sheet1"  # change if your sheet name is different

# Interactive Brokers API connection settings.
# HOST and PORT must match what is configured in TWS / IB Gateway.
IB_HOST = "127.0.0.1"
IB_PORT = 7496  # 7496 = live trading, 7497 = paper by default
IB_CLIENT_ID = 1

# Account to send orders to (empty string = IB default).
# Keeping this as a single constant makes it easy to switch accounts later.
IB_ACCOUNT = "U24159961"

# Trading defaults
DEFAULT_EXCHANGE = "SMART"
DEFAULT_CURRENCY = "USD"

# Safety switch: when True, the script will ONLY print planned orders and never send them.
# Flip to False only when you are confident everything is configured correctly.
DRY_RUN = False


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def normalize_direction(direction_cell) -> str:
    """
    Normalize the direction text coming from Excel.

    We only care about 'long' and 'short', but we keep any other text
    as-is so it is easy to spot invalid values in the log output.
    """
    if direction_cell is None:
        return ""
    s = str(direction_cell).strip().lower()
    if s in ("long", "short"):
        return s
    return s


def read_live_trade_info() -> List[Tuple[str, str, int]]:
    """
    Read Live_Trade_Info.xlsx and build a clean list of trades.

    - Expects row 1 headers: A=Ticker, B=Direction, C=Share Size.
    - Converts direction ('long'/'short') to IB actions ('BUY'/'SELL').
    - Validates that share size is a positive integer.
    - Skips any rows that are incomplete or invalid while printing a reason.
    """
    if not os.path.exists(LIVE_INFO_FILE):
        raise FileNotFoundError(f"Live trade info file not found: {LIVE_INFO_FILE}")

    wb = load_workbook(LIVE_INFO_FILE, data_only=True)
    if LIVE_INFO_SHEET in wb.sheetnames:
        ws = wb[LIVE_INFO_SHEET]
    else:
        ws = wb.active

    trades: List[Tuple[str, str, int]] = []

    # Expect header in row 1: A=Ticker, B=Direction, C=Share Size
    for row in range(2, ws.max_row + 1):
        ticker_cell = ws.cell(row=row, column=1).value
        direction_cell = ws.cell(row=row, column=2).value
        size_cell = ws.cell(row=row, column=3).value

        if ticker_cell is None or str(ticker_cell).strip() == "":
            continue

        ticker = str(ticker_cell).strip().upper()
        direction_norm = normalize_direction(direction_cell)

        if direction_norm not in ("long", "short"):
            print(f"Row {row}: invalid direction '{direction_cell}' for ticker {ticker}; skipping.")
            continue

        try:
            size = int(size_cell)
        except (TypeError, ValueError):
            print(f"Row {row}: invalid share size '{size_cell}' for ticker {ticker}; skipping.")
            continue

        if size <= 0:
            print(f"Row {row}: non-positive share size {size} for ticker {ticker}; skipping.")
            continue

        action = "BUY" if direction_norm == "long" else "SELL"
        trades.append((ticker, action, size))

    return trades


def connect_ib() -> IB:
    """
    Connect to Interactive Brokers using ib_insync with the configured host/port.

    Any import or connection issues are surfaced as clear error messages so they
    are easy to diagnose from the console output.
    """
    if IB is None:
        raise ImportError(
            "Could not import ib_insync or its dependencies. "
            "Details:\n"
            f"    {IB_IMPORT_ERROR}\n\n"
            "Try reinstalling with:\n"
            "    python -m pip install --upgrade ib_insync eventkit nest-asyncio numpy"
        )

    ib = IB()
    print(f"Connecting to IB at {IB_HOST}:{IB_PORT} with clientId={IB_CLIENT_ID} ...")
    ib.connect(IB_HOST, IB_PORT, clientId=IB_CLIENT_ID)
    print("Connected to IB.")
    return ib


def place_trades_ib(ib: IB, trades: List[Tuple[str, str, int]]) -> None:
    """
    Print a summary of planned trades and, when DRY_RUN is False, send
    Market-On-Close (MOC) orders for each one.

    Orders are created as:
      - action: BUY for long, SELL for short
      - type:  MOC (market-on-close)
      - tif:   DAY (good for this trading day only)
      - account: IB_ACCOUNT when set
    """
    if not trades:
        print("No trades to place.")
        return

    print("\nPlanned trades:")
    for ticker, action, size in trades:
        print(f"  {action} {size} {ticker}")

    if DRY_RUN:
        print("\nDRY_RUN is True: no orders will be sent. "
              "Set DRY_RUN = False at the top of this script to send live orders.")
        return

    print("\nPlacing market-on-close (MOC) orders...")
    for ticker, action, size in trades:
        contract = Stock(ticker, DEFAULT_EXCHANGE, DEFAULT_CURRENCY)
        # MOC order via generic Order: type 'MOC', market-on-close
        order = Order(
            action=action,
            orderType="MOC",
            totalQuantity=size,
            tif="DAY",  # time in force: DAY
        )
        if IB_ACCOUNT:
            order.account = IB_ACCOUNT

        trade = ib.placeOrder(contract, order)
        print(f"Submitted {action} {size} {ticker}, orderId={trade.order.orderId}")

    # Give IB a moment to process, then print final statuses
    ib.sleep(2)
    print("\nOrder statuses:")
    for t in ib.trades():
        status = t.orderStatus.status
        filled = t.orderStatus.filled
        remaining = t.orderStatus.remaining
        print(
            f"  orderId={t.order.orderId} status={status} "
            f"filled={filled} remaining={remaining}"
        )


def main():
    """
    Orchestrate the full workflow:
      1) Read and validate trades from Live_Trade_Info.xlsx.
      2) Connect to IB via TWS / Gateway.
      3) Either print the trades only (DRY_RUN) or submit MOC orders.
      4) Cleanly disconnect from IB regardless of success/failure.
    """
    try:
        trades = read_live_trade_info()
    except FileNotFoundError as e:
        print(e)
        return

    if not trades:
        print("No valid trades found in Live_Trade_Info; nothing to do.")
        return

    try:
        ib = connect_ib()
    except Exception as e:
        print(f"Failed to connect to Interactive Brokers: {e}")
        return

    try:
        place_trades_ib(ib, trades)
    finally:
        print("Disconnecting from IB...")
        ib.disconnect()
        print("Disconnected.")


if __name__ == "__main__":
    main()
