"""
Obtain Live Trade Info

Reads tickers and trade parameters from the Earnings workbook and writes them
to a Live_Trade_Info workbook that will later be used to send live orders.

Source layout (! -- Latest Earnings Document.xlsx):
- Column A: Ticker
- Column Y: Trade Direction ("long" or "short")
- Column Z: Share Size

Target layout (Live_Trade_Info.xlsx):
- Row 1 headers: A1=Ticker, B1=Direction, C1=Share Size
- Rows 2+ will be overwritten each run with the selected rows.
- Column D (Exit Type) is filled by Exit_Live_Trades_IBKR.py from Earnings column R.
"""

import os
import warnings

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

warnings.filterwarnings("ignore", message=".*Unknown extension.*", category=UserWarning)
warnings.filterwarnings("ignore", message=".*Conditional Formatting extension.*", category=UserWarning)

# ---------------------------------------------------------------------------
# Paths: Excel files live in repo root (Trading_Algo); script is in Old Scripts
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR = os.path.dirname(os.path.dirname(_SCRIPT_DIR))

# ---------------------------------------------------------------------------
# Configuration — update paths/sheet/columns as needed
# ---------------------------------------------------------------------------
# Path to the Earnings workbook (e.g. quarterly earnings / trade setup file).
source_file = os.path.join(_BASE_DIR, "! -- Latest Earnings Document.xlsx")
# Sheet name that contains the ticker list and trade direction/size.
source_sheet = "Trades"  # Sheet that contains the trade setup
# Row index where column headers live; data rows start below this.
header_row = 3  # Row where column headers live (data starts below this)

# Column letters in the Earnings workbook for ticker, direction, and share size.
ticker_column_letter = "A"
direction_column_letter = "Y"
size_column_letter = "Z"

# Output workbook path. This file is read by Send_Live_Trades_IBKR.py to place orders.
output_file = os.path.join(_BASE_DIR, "Live_Trade_Info.xlsx")
# Sheet name in the output workbook; change if your Live_Trade_Info uses a different sheet.
output_sheet = "Sheet1"  # Use the first sheet by default; adjust if needed


# ---------------------------------------------------------------------------
# Helpers: normalize values read from Excel
# ---------------------------------------------------------------------------

def normalize_ticker(cell_value):
    """
    Normalize ticker text from a cell: strip whitespace and uppercase.
    Dots are preserved (e.g. BRK.B) since this file feeds trading and IB uses the same format.
    """
    if cell_value is None:
        return None
    return str(cell_value).strip().upper()


def normalize_direction(cell_value):
    """
    Normalize trade direction to lowercase 'long' or 'short'.
    Returns the raw string if not long/short so invalid values are visible in logs.
    """
    if cell_value is None:
        return None
    s = str(cell_value).strip().lower()
    if s in ("long", "short"):
        return s
    return s  # return whatever is there so you can spot issues


def main():
    """
    Full workflow: prompt for row range, read trades from Earnings, write to Live_Trade_Info.
    Skips rows with missing ticker, direction, or share size and prints a message for each.
    """
    # -----------------------------------------------------------------------
    # Check source file exists before prompting for input
    # -----------------------------------------------------------------------
    if not os.path.exists(source_file):
        print(f"Source Earnings file not found:\n  {source_file}")
        print("\nEnsure your Earnings workbook is in the Trading_Algo folder (parent of the scripts folder).")
        return

    # -----------------------------------------------------------------------
    # Ask which row range to pull from (same idea as Update_Closing_Prices row range)
    # -----------------------------------------------------------------------
    try:
        start_row = int(
            input(
                f"Enter starting row number to process "
                f"(first data row is typically {header_row + 1}): "
            )
        )
        end_row = int(input("Enter ending row number to process: "))
    except ValueError:
        print("Row numbers must be integers.")
        return

    if end_row < start_row:
        print("Ending row must be >= starting row.")
        return

    # -----------------------------------------------------------------------
    # Load source workbook and sheet (data_only=True so formulas show values)
    # -----------------------------------------------------------------------
    wb_source = load_workbook(source_file, data_only=True)
    try:
        ws_source = wb_source[source_sheet]
    except KeyError:
        print(f"Worksheet '{source_sheet}' not found in {source_file}.")
        return

    # -----------------------------------------------------------------------
    # Collect trade rows from the Earnings sheet (A, Y, Z)
    # -----------------------------------------------------------------------
    trades = []  # list of (ticker, direction, size)

    for row in range(start_row, end_row + 1):
        raw_ticker = ws_source[f"{ticker_column_letter}{row}"].value
        raw_direction = ws_source[f"{direction_column_letter}{row}"].value
        raw_size = ws_source[f"{size_column_letter}{row}"].value

        ticker = normalize_ticker(raw_ticker)
        direction = normalize_direction(raw_direction)
        size = raw_size  # keep as-is; numeric or formula result is fine

        # Skip rows without a ticker
        if not ticker:
            continue

        # Require direction and share size; skip and warn so user can fix the sheet
        if not direction:
            print(f"Row {row}: missing trade direction for ticker {ticker}; skipping.")
            continue
        if size is None or str(size).strip() == "":
            print(f"Row {row}: missing share size for ticker {ticker}; skipping.")
            continue

        trades.append((ticker, direction, size))

    if not trades:
        print("No valid trades found in the specified row range.")
        return

    print(f"Collected {len(trades)} trade(s) from Earnings file.")

    # -----------------------------------------------------------------------
    # Load or create the Live_Trade_Info workbook
    # -----------------------------------------------------------------------
    if os.path.exists(output_file):
        wb_output = load_workbook(output_file)
        if output_sheet in wb_output.sheetnames:
            ws_output = wb_output[output_sheet]
        else:
            ws_output = wb_output.active
    else:
        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = output_sheet

    # -----------------------------------------------------------------------
    # Ensure header row exists (row 1: Ticker, Direction, Share Size)
    # -----------------------------------------------------------------------
    if ws_output.max_row < 1:
        ws_output.append(["Ticker", "Direction", "Share Size"])
    else:
        # If row 1 is empty, set the headers so Send_Live_Trades_IBKR finds them
        if (
            ws_output["A1"].value is None
            and ws_output["B1"].value is None
            and ws_output["C1"].value is None
        ):
            ws_output["A1"] = "Ticker"
            ws_output["B1"] = "Direction"
            ws_output["C1"] = "Share Size"

    # -----------------------------------------------------------------------
    # Clear any existing data below the header (rows 2+) so we write a fresh set
    # -----------------------------------------------------------------------
    if ws_output.max_row > 1:
        ws_output.delete_rows(2, ws_output.max_row - 1)

    # -----------------------------------------------------------------------
    # Write one row per trade: A=ticker, B=direction, C=share size
    # -----------------------------------------------------------------------
    for ticker, direction, size in trades:
        ws_output.append([ticker, direction, size])

    # Left-align headers and data for a clean look (A, B, C)
    left_align = Alignment(horizontal="left")
    for col in range(1, 4):  # A, B, C
        ws_output.cell(row=1, column=col).alignment = left_align
    for row in range(2, ws_output.max_row + 1):
        for col in range(1, 4):
            ws_output.cell(row=row, column=col).alignment = left_align

    # -----------------------------------------------------------------------
    # Save the workbook and confirm
    # -----------------------------------------------------------------------
    wb_output.save(output_file)
    print(
        f"Successfully wrote {len(trades)} trade(s) to '{output_file}' "
        f"in sheet '{ws_output.title}'."
    )


if __name__ == "__main__":
    main()
