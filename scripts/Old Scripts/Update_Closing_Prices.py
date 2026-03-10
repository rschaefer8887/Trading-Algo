"""
Update Closing Prices (date-column layout)

Reads tickers from the Earnings workbook (given row range) and from the existing
Closing_Prices file. Adds or updates a column for a specific date (chosen by the
user) and fills closing prices for all tickers already in the file. Any ticker
in the Earnings range that is not yet in Closing_Prices is added at the bottom
of column A; for those new tickers, only the chosen date's column gets a price
(older date columns stay blank).

Layout: Column A = Ticker, B1 = first date, C1 = second date, ... (one column per
run date). Rows 2+ = one ticker per row with closing prices under each date.
"""

import os
import warnings
from datetime import datetime, timedelta

import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

warnings.filterwarnings("ignore", message=".*Unknown extension.*", category=UserWarning)
warnings.filterwarnings("ignore", message=".*Conditional Formatting extension.*", category=UserWarning)

# ---------------------------------------------------------------------------
# Paths: Excel files live in repo root (Trading Algo); script is in Old Scripts
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR = os.path.dirname(os.path.dirname(_SCRIPT_DIR))

# ---------------------------------------------------------------------------
# Configuration — update source file path when you refresh your earnings file
# ---------------------------------------------------------------------------
source_file = os.path.join(_BASE_DIR, "! -- Latest Earnings Document.xlsx")
source_sheet = "Trades"  # Sheet that contains the ticker list
header_row = 3  # Row index where column headers live (tickers start below this)
ticker_column_letter = "A"  # Column containing ticker symbols

output_file = os.path.join(_BASE_DIR, "Closing_Prices.xlsx")
output_sheet = "Prices"  # Name of the sheet in the output workbook

# ---------------------------------------------------------------------------
# Check that the source file exists before continuing
# ---------------------------------------------------------------------------
if not os.path.exists(source_file):
    print(f"Source file not found:\n  {source_file}")
    print("\nEnsure your Earnings workbook is in the Trading Algo folder (parent of the scripts folder).")
    exit(1)

# ---------------------------------------------------------------------------
# Get the row range (used to discover new tickers from the Earnings file)
# ---------------------------------------------------------------------------
start_row = int(input(f"Enter starting row number to process (first ticker is row {header_row+1}): "))
end_row = int(input("Enter ending row number to process: "))

if end_row < start_row:
    print("Ending row must be >= starting row.")
    exit()

# ---------------------------------------------------------------------------
# Get the target date to fetch closing prices for
# ---------------------------------------------------------------------------
date_input = input(
    "Enter date to fetch closing prices for "
    "(e.g. YYYY-MM-DD or MM/DD/YY): "
)

# ---------------------------------------------------------------------------
# Load the source workbook and read ticker symbols from the chosen range
# ---------------------------------------------------------------------------
wb_source = load_workbook(source_file, data_only=True)
try:
    ws_source = wb_source[source_sheet]
except KeyError:
    print(f"Worksheet '{source_sheet}' not found in {source_file}.")
    exit()

def normalize_ticker(cell_value):
    if cell_value is None:
        return None
    return str(cell_value).strip().upper().replace(".", "-")

def parse_date_to_canonical(val):
    """Parse a header cell (datetime or string) to YYYY-MM-DD for internal use."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # fallback: use as-is

def format_date_display(canonical_str):
    """Format YYYY-MM-DD as MM/DD/YY for row 1 headers."""
    try:
        return datetime.strptime(canonical_str, "%Y-%m-%d").strftime("%m/%d/%y")
    except (ValueError, TypeError):
        return canonical_str

# ---------------------------------------------------------------------------
# Normalize the requested date into canonical YYYY-MM-DD format
# ---------------------------------------------------------------------------
target_date = parse_date_to_canonical(date_input)
if not target_date:
    print("Invalid date format. Please use YYYY-MM-DD or MM/DD/YY.")
    exit()

earnings_tickers = []
for row in range(start_row, end_row + 1):
    cell = ws_source[f"{ticker_column_letter}{row}"].value
    t = normalize_ticker(cell)
    if t:
        earnings_tickers.append(t)

# Deduplicate while preserving order (first occurrence wins)
seen = set()
earnings_tickers = [t for t in earnings_tickers if t not in seen and not seen.add(t)]

# ---------------------------------------------------------------------------
# Load existing Closing_Prices (if any): tickers in A, date headers in row 1, data grid
# ---------------------------------------------------------------------------
existing_tickers = []
existing_dates = []
data = {}  # (ticker, date_str) -> value (number or None)

if os.path.exists(output_file):
    wb_out = load_workbook(output_file, data_only=True)
    if output_sheet in wb_out.sheetnames:
        ws = wb_out[output_sheet]
    else:
        ws = wb_out.active
    # Row 1: A1 = "Ticker" or similar, B1 = first date, C1 = second, ...
    for col in range(2, ws.max_column + 1):
        val = ws.cell(1, col).value
        canonical = parse_date_to_canonical(val)
        if canonical:
            existing_dates.append(canonical)
    # Column A from row 2: ticker list
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val is not None and str(val).strip():
            existing_tickers.append(normalize_ticker(val))
    # Data grid: (ticker, date) -> value
    for r, ticker in enumerate(existing_tickers):
        row = r + 2
        for c, date_str in enumerate(existing_dates):
            col = c + 2
            cell_val = ws.cell(row, col).value
            data[(ticker, date_str)] = cell_val
    wb_out.close()

# ---------------------------------------------------------------------------
# Merge: existing tickers stay; new tickers (in Earnings but not in file) go at bottom
# ---------------------------------------------------------------------------
existing_set = set(existing_tickers)
new_tickers = [t for t in earnings_tickers if t not in existing_set]
final_tickers = existing_tickers + new_tickers

if not final_tickers:
    print("No tickers to update (none in Earnings range and no existing Closing_Prices rows).")
    exit()

if new_tickers:
    print(f"New tickers to add: {new_tickers}")
print(f"Total tickers: {len(final_tickers)} (existing: {len(existing_tickers)}, new: {len(new_tickers)})")

# ---------------------------------------------------------------------------
# Target date column (selected by the user)
# ---------------------------------------------------------------------------
if target_date not in existing_dates:
    existing_dates.append(target_date)
dates = existing_dates

# ---------------------------------------------------------------------------
# Fetch closing prices for the chosen date for all final_tickers
# ---------------------------------------------------------------------------
target_dt = datetime.strptime(target_date, "%Y-%m-%d")
next_dt = target_dt + timedelta(days=1)

ydata = yf.download(
    final_tickers,
    start=target_dt.strftime("%Y-%m-%d"),
    end=next_dt.strftime("%Y-%m-%d"),
    group_by="ticker",
    auto_adjust=True,
    threads=True,
)

if ydata.empty:
    print("No data returned from yfinance for the selected date!")
    exit()

today_prices = {}
for ticker in final_tickers:
    try:
        if len(final_tickers) > 1:
            price = ydata[ticker]["Close"].iloc[-1]
        else:
            price = ydata["Close"].iloc[-1]
        today_prices[ticker] = round(float(price), 2) if price is not None else None
    except Exception as e:
        print(f"Could not get price for {ticker}: {e}")
        today_prices[ticker] = None

# Update data with the chosen date's prices
for ticker in final_tickers:
    data[(ticker, target_date)] = today_prices.get(ticker)

# ---------------------------------------------------------------------------
# Write the output workbook
# ---------------------------------------------------------------------------
wb_output = Workbook()
ws_output = wb_output.active
ws_output.title = output_sheet

# Row 1: Ticker, then date headers (MM/DD/YY format)
ws_output.append(["Ticker"] + [format_date_display(d) for d in dates])

# Rows 2+: one row per ticker, then value for each date (blank if missing)
for ticker in final_tickers:
    row_vals = [ticker] + [
        round(v, 2) if isinstance(v, (int, float)) else v
        for v in [data.get((ticker, d), "") for d in dates]
    ]
    ws_output.append(row_vals)

# Left-align header row (Ticker + date columns) and ticker column (column A)
left_align = Alignment(horizontal="left")
for col in range(1, len(dates) + 2):  # Ticker + one per date
    ws_output.cell(row=1, column=col).alignment = left_align
for row in range(2, len(final_tickers) + 2):
    ws_output.cell(row=row, column=1).alignment = left_align
# Left-align all date/value cells in rows 2+ for a consistent look
for row in range(2, len(final_tickers) + 2):
    for col in range(2, len(dates) + 2):
        ws_output.cell(row=row, column=col).alignment = left_align

wb_output.save(output_file)
print(f"Successfully updated '{output_file}' with {len(dates)} date column(s) and {len(final_tickers)} tickers.")
