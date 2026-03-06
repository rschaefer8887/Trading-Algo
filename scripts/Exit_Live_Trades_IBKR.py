"""
Exit Live Trades to Interactive Brokers based on Live_Trade_Info.xlsx

Reads the same tickers and share sizes as your entry orders (from Live_Trade_Info.xlsx)
and places exit orders in the **opposite** direction to close the positions:
- Entry was LONG  (bought)  → Exit: SELL (same size) to close the long
- Entry was SHORT (sold)     → Exit: BUY  (same size) to cover the short

Exit order type is read from column R in the Latest Earnings file, then written to
column D (Exit Type) of Live_Trade_Info so you can verify before sending:
- "MOC"  → Market-on-close order
- "Open" → Market order (executes during the session)

After updating Live_Trade_Info column D, the script prompts whether to send live exit
orders. Run this after your entry orders have filled and you want to exit.

Prerequisites:
- IBKR Desktop / TWS or IB Gateway running with API enabled.
- Python package 'ib_insync' installed:
    pip install ib_insync
"""

import os
import asyncio
import warnings
from typing import List, Tuple

from openpyxl import load_workbook

warnings.filterwarnings("ignore", message=".*Unknown extension.*", category=UserWarning)
warnings.filterwarnings("ignore", message=".*Conditional Formatting extension.*", category=UserWarning)

# Ensure there is an asyncio event loop for ib_insync/eventkit on newer Python versions.
try:
    asyncio.get_event_loop()
except RuntimeError:
    asyncio.set_event_loop(asyncio.new_event_loop())

IB_IMPORT_ERROR = None
try:
    from ib_insync import IB, Stock, Order
except Exception as e:
    IB = None  # type: ignore[assignment]
    IB_IMPORT_ERROR = e


# ---------------------------------------------------------------------------
# Paths: Excel files live in the Trading Algo folder (parent of this script)
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR = os.path.dirname(_SCRIPT_DIR)

# ---------------------------------------------------------------------------
# Configuration — must match your entry script / TWS
# ---------------------------------------------------------------------------
LIVE_INFO_FILE = os.path.join(_BASE_DIR, "Live_Trade_Info.xlsx")
LIVE_INFO_SHEET = "Sheet1"

# Earnings file: used to read exit type (column R) per ticker (column A)
EARNINGS_FILE = os.path.join(_BASE_DIR, "! -- Latest Earnings Document.xlsx")
EARNINGS_SHEET = "Trades"
EARNINGS_HEADER_ROW = 3  # data rows start below this
EARNINGS_TICKER_COL = "A"
EARNINGS_EXIT_TYPE_COL = "R"

IB_HOST = "127.0.0.1"
IB_PORT = 7496  # 7496 = live, 7497 = paper
IB_CLIENT_ID = 2  # Use a different client ID than entry script so both can run if needed
IB_ACCOUNT = "U1867866"

DEFAULT_EXCHANGE = "SMART"
DEFAULT_CURRENCY = "USD"

# When True, only print planned exit orders; do not send them.
DRY_RUN = False


def normalize_direction(direction_cell) -> str:
    if direction_cell is None:
        return ""
    s = str(direction_cell).strip().lower()
    if s in ("long", "short"):
        return s
    return s


def normalize_exit_type(cell_value):
    """Return (order_type, display): 'Open' -> ('MKT', 'Open'), else ('MOC', 'MOC')."""
    if cell_value is None or not str(cell_value).strip():
        return "MOC", "MOC"
    s = str(cell_value).strip()
    s_lower = s.lower()
    if s_lower == "open":
        return "MKT", "Open"
    return "MOC", "MOC"


def _build_ticker_to_exit_type() -> dict:
    """Read Earnings file column A and R; return dict ticker_upper -> (order_type, display)."""
    if not os.path.exists(EARNINGS_FILE):
        return {}
    wb = load_workbook(EARNINGS_FILE, data_only=True)
    try:
        ws = wb[EARNINGS_SHEET]
    except KeyError:
        return {}
    result = {}
    for row in range(EARNINGS_HEADER_ROW + 1, ws.max_row + 1):
        ticker_cell = ws[f"{EARNINGS_TICKER_COL}{row}"].value
        exit_cell = ws[f"{EARNINGS_EXIT_TYPE_COL}{row}"].value
        if ticker_cell is None or not str(ticker_cell).strip():
            continue
        ticker_key = str(ticker_cell).strip().upper()
        result[ticker_key] = normalize_exit_type(exit_cell)
    return result


def read_exit_trade_info():
    """
    Read Live_Trade_Info.xlsx (A, B, C only) and Earnings (column R) for exit type.
    Build exit orders: long → SELL, short → BUY, same size; order type from Earnings R.
    Returns (exits, display_values): exits = [(ticker, action, size, order_type), ...],
    display_values = ["MOC" or "Open", ...] for writing to Live_Trade_Info column D.
    """
    if not os.path.exists(LIVE_INFO_FILE):
        raise FileNotFoundError(f"Live trade info file not found: {LIVE_INFO_FILE}")

    ticker_to_exit = _build_ticker_to_exit_type()

    wb = load_workbook(LIVE_INFO_FILE, data_only=True)
    if LIVE_INFO_SHEET in wb.sheetnames:
        ws = wb[LIVE_INFO_SHEET]
    else:
        ws = wb.active

    exits: List[Tuple[str, str, int, str]] = []
    display_values: List[str] = []

    for row in range(2, ws.max_row + 1):
        ticker_cell = ws.cell(row=row, column=1).value
        direction_cell = ws.cell(row=row, column=2).value
        size_cell = ws.cell(row=row, column=3).value

        if ticker_cell is None or str(ticker_cell).strip() == "":
            continue

        ticker = str(ticker_cell).strip().upper()
        direction_norm = normalize_direction(direction_cell)

        if direction_norm not in ("long", "short"):
            print(f"Row {row}: invalid direction '{direction_cell}' for ticker {ticker}; skipping.")
            continue

        try:
            size = int(size_cell)
        except (TypeError, ValueError):
            print(f"Row {row}: invalid share size '{size_cell}' for ticker {ticker}; skipping.")
            continue

        if size <= 0:
            print(f"Row {row}: non-positive share size {size} for ticker {ticker}; skipping.")
            continue

        order_type, display = ticker_to_exit.get(ticker, ("MOC", "MOC"))
        action = "SELL" if direction_norm == "long" else "BUY"
        exits.append((ticker, action, size, order_type))
        display_values.append(display)

    return exits, display_values


def write_exit_types_to_live_info(display_values: List[str]) -> None:
    """Write Exit Type to column D of Live_Trade_Info (D1=header, D2+=values). Leave A,B,C unchanged."""
    wb = load_workbook(LIVE_INFO_FILE)  # data_only=False so we can write
    if LIVE_INFO_SHEET in wb.sheetnames:
        ws = wb[LIVE_INFO_SHEET]
    else:
        ws = wb.active
    ws.cell(row=1, column=4).value = "Exit Type"
    for i, val in enumerate(display_values):
        ws.cell(row=i + 2, column=4).value = val
    wb.save(LIVE_INFO_FILE)
    print(f"Updated Exit Type (column D) in '{LIVE_INFO_FILE}'.")


def connect_ib() -> IB:
    if IB is None:
        raise ImportError(
            "Could not import ib_insync or its dependencies. "
            f"Details:\n    {IB_IMPORT_ERROR}\n\n"
            "Try: python -m pip install --upgrade ib_insync eventkit nest-asyncio numpy"
        )
    ib = IB()
    print(f"Connecting to IB at {IB_HOST}:{IB_PORT} with clientId={IB_CLIENT_ID} ...")
    ib.connect(IB_HOST, IB_PORT, clientId=IB_CLIENT_ID)
    print("Connected to IB.")
    return ib


def place_exit_orders_ib(ib: IB, exits: List[Tuple[str, str, int, str]]) -> None:
    """Place exit orders (MOC or MKT per row) to close/cover positions."""
    if not exits:
        print("No exit orders to place.")
        return

    print("\nPlanned exit orders (close/cover):")
    for ticker, action, size, order_type in exits:
        print(f"  {action} {size} {ticker}  [{order_type}]")

    if DRY_RUN:
        print("\nDRY_RUN is True: no orders will be sent. Set DRY_RUN = False to send exit orders.")
        return

    print("\nPlacing exit orders...")
    for ticker, action, size, order_type in exits:
        contract = Stock(ticker, DEFAULT_EXCHANGE, DEFAULT_CURRENCY)
        order = Order(
            action=action,
            orderType=order_type,  # "MOC" or "MKT"
            totalQuantity=size,
            tif="DAY",
        )
        if IB_ACCOUNT:
            order.account = IB_ACCOUNT
        trade = ib.placeOrder(contract, order)
        print(f"Submitted {action} {size} {ticker} ({order_type}), orderId={trade.order.orderId}")

    ib.sleep(2)
    print("\nOrder statuses:")
    for t in ib.trades():
        print(
            f"  orderId={t.order.orderId} status={t.orderStatus.status} "
            f"filled={t.orderStatus.filled} remaining={t.orderStatus.remaining}"
        )


def main():
    try:
        exits, display_values = read_exit_trade_info()
    except FileNotFoundError as e:
        print(e)
        return

    if not exits:
        print("No valid rows in Live_Trade_Info; nothing to exit.")
        return

    # Write exit types from Earnings (column R) into Live_Trade_Info column D
    write_exit_types_to_live_info(display_values)

    # Prompt before sending live orders
    reply = input("\nSend live exit orders? (y/n): ").strip().lower()
    if reply not in ("y", "yes"):
        print("Exiting without sending orders.")
        return

    try:
        ib = connect_ib()
    except Exception as e:
        print(f"Failed to connect to Interactive Brokers: {e}")
        return

    try:
        place_exit_orders_ib(ib, exits)
    finally:
        print("Disconnecting from IB...")
        ib.disconnect()
        print("Disconnected.")


if __name__ == "__main__":
    main()
