"""
Get Closing Prices

Reads a list of stock tickers from an Excel "Earnings" workbook (Trades sheet),
fetches the latest closing price for each via Yahoo Finance (yfinance), and
writes the results to a separate Excel file (Closing_Prices.xlsx).

Update the source Earnings file path/sheet as needed (e.g. once per quarter).
"""

import yfinance as yf
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# ---------------------------------------------------------------------------
# Paths: Excel files live in the Trading Algo folder (project root, two levels up from this script)
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR = os.path.dirname(os.path.dirname(_SCRIPT_DIR))

# ---------------------------------------------------------------------------
# Configuration — update source file path when you refresh your earnings file
# ---------------------------------------------------------------------------
source_file = os.path.join(_BASE_DIR, "! -- Latest Earnings Document.xlsx")
source_sheet = "Trades"  # Sheet that contains the ticker list
header_row = 3  # Row index where column headers live (tickers start below this)
ticker_column_letter = "A"  # Column containing ticker symbols

output_file = os.path.join(_BASE_DIR, "Closing_Prices_One_Time.xlsx")
output_sheet = "Prices"  # Name of the sheet in the output workbook

# ---------------------------------------------------------------------------
# Get the row range to process (user specifies which rows have tickers)
# ---------------------------------------------------------------------------
start_row = int(input(f"Enter starting row number to process (first ticker is row {header_row+1}): "))
end_row = int(input("Enter ending row number to process: "))

if end_row < start_row:
    print("Ending row must be >= starting row.")
    exit()

# ---------------------------------------------------------------------------
# Load the source workbook and read ticker symbols from the chosen column
# ---------------------------------------------------------------------------
wb_source = load_workbook(source_file, data_only=True)  # data_only=True gives values, not formulas
try:
    ws_source = wb_source[source_sheet]
except KeyError:
    print(f"Worksheet '{source_sheet}' not found in {source_file}.")
    exit()

# Build list of tickers from column A in the given row range (skip empty cells)
tickers = []
for row in range(start_row, end_row + 1):
    cell = ws_source[f"{ticker_column_letter}{row}"].value
    if cell is None:
        continue
    # Normalize: strip spaces, uppercase, and convert dots to hyphens for Yahoo (e.g. BRK.B -> BRK-B)
    ticker = str(cell).strip().upper().replace(".", "-")
    tickers.append(ticker)

if not tickers:
    print("No tickers found in the specified rows.")
    exit()

print(f"Found {len(tickers)} tickers: {tickers}")

# ---------------------------------------------------------------------------
# Fetch latest closing prices from Yahoo Finance (1-day period = latest close)
# ---------------------------------------------------------------------------
# group_by='ticker' keeps each ticker in its own column when multiple tickers
# auto_adjust=True uses adjusted close; threads=True speeds up multi-ticker download
data = yf.download(tickers, period="1d", group_by='ticker', auto_adjust=True, threads=True)

if data.empty:
    print("No data returned from yfinance!")
    exit()

# ---------------------------------------------------------------------------
# Prepare the output workbook: create new or clear existing file
# ---------------------------------------------------------------------------
if os.path.exists(output_file):
    wb_output = load_workbook(output_file)
    ws_output = wb_output.active
    ws_output.title = output_sheet
    # Remove all existing rows so we write a fresh set of results
    ws_output.delete_rows(1, ws_output.max_row)
else:
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = output_sheet

# Column headers for the output sheet
ws_output.append(["Ticker", "Date", "Closing Price"])

today = datetime.today().strftime("%Y-%m-%d")

# ---------------------------------------------------------------------------
# Write one row per ticker: symbol, date, and closing price (or None if fetch failed)
# ---------------------------------------------------------------------------
for ticker in tickers:
    try:
        # yfinance returns different DataFrame shapes: multi-level columns for 2+ tickers, flat for 1
        if len(tickers) > 1:
            price = data[ticker]['Close'].iloc[-1]
        else:
            price = data['Close'].iloc[-1]
        ws_output.append([ticker, today, price])
    except Exception as e:
        print(f"Could not get price for {ticker}: {e}")
        ws_output.append([ticker, today, None])

# ---------------------------------------------------------------------------
# Save the output workbook and confirm
# ---------------------------------------------------------------------------
wb_output.save(output_file)
print(f"Successfully wrote {len(tickers)} tickers to '{output_file}'")
