"""
Stage Trades Auto — Build Live_Trade_Info from Latest Earnings using column O flags

Reads the Latest Earnings workbook (Trades sheet). Column O contains flags:
  - "T" → single-day: write those rows to sheet "Daily_Trades".
  - "1" → Monday, "2" → Tuesday, "3" → Wednesday, "4" → Thursday, "5" → Friday.
  - If any 1/2/3/4 are found, use only weekday rows (ignore T) and print
    "No clear single day range, writing weekday trades."

For each row: ticker from A, direction from Y, share size from Z,
IBKR Exit from AB, ToS Exit from AA.

Writes to Live_Trade_Info.xlsx (same layout for Daily_Trades and weekday sheets):
  - Row 1: A1=Ticker, B1=Direction, C1=Share Size, D1=IBKR Exit, E1=ToS Exit
  - Rows 2+: one row per trade.
"""

import os
import sys
import warnings
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

warnings.filterwarnings("ignore", message=".*Unknown extension.*", category=UserWarning)
warnings.filterwarnings("ignore", message=".*Conditional Formatting extension.*", category=UserWarning)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR = os.path.dirname(_SCRIPT_DIR)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SOURCE_FILE = os.path.join(_BASE_DIR, "! -- Latest Earnings Document.xlsx")
SOURCE_SHEET = "Trades"
HEADER_ROW = 3  # Data starts the row after headers (e.g. row 4)
COL_FLAG = "O"
COL_TICKER = "A"
COL_DIRECTION = "Y"
COL_SIZE = "Z"
COL_IBKR_EXIT = "AB"   # Source for Live_Trade_Info column D
COL_TOS_EXIT = "AA"    # Source for Live_Trade_Info column E

OUTPUT_FILE = os.path.join(_BASE_DIR, "Live_Trade_Info.xlsx")
DAILY_SHEET = "Daily_Trades"
WEEKDAY_SHEETS = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
HEADER_D1 = "IBKR Exit"
HEADER_E1 = "ToS Exit"

# Column O: 1 -> Monday, 2 -> Tuesday, 3 -> Wednesday, 4 -> Thursday, 5 -> Friday
_FLAG_TO_WEEKDAY = {1: "Monday", 2: "Tuesday", 3: "Wednesday", 4: "Thursday", 5: "Friday"}


def _parse_flag(cell_value: Any) -> Optional[Any]:
    """
    Parse column O. Returns "T", 1, 2, 3, 4, 5, or None (skip).
    Accepts string or int for numbers (e.g. "1" or 1 -> 1). 5 = Friday.
    """
    if cell_value is None:
        return None
    s = str(cell_value).strip().upper()
    if s == "T":
        return "T"
    try:
        n = int(cell_value) if isinstance(cell_value, (int, float)) else int(s)
        if 1 <= n <= 5:
            return n
    except (ValueError, TypeError):
        pass
    return None


def _normalize_ticker(cell_value: Any) -> Optional[str]:
    if cell_value is None:
        return None
    return str(cell_value).strip().upper()


def _normalize_direction(cell_value: Any) -> Optional[str]:
    if cell_value is None:
        return None
    return str(cell_value).strip().lower()


def _cell_to_str(cell_value: Any) -> str:
    """Return cell value as string for exit-type columns; empty string if None."""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


# Trade tuple: (ticker, direction, size, ibkr_exit, tos_exit)
TradeRow = Tuple[str, str, Any, str, str]


def _write_trades_to_sheet(ws: Worksheet, trades: List[TradeRow]) -> None:
    """Write header row and trade rows to sheet (same format as Daily_Trades)."""
    ws["A1"] = "Ticker"
    ws["B1"] = "Direction"
    ws["C1"] = "Share Size"
    ws["D1"] = HEADER_D1
    ws["E1"] = HEADER_E1
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for ticker, direction, size, ibkr_exit, tos_exit in trades:
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=ticker)
        ws.cell(row=next_row, column=2, value=direction)
        ws.cell(row=next_row, column=3, value=size)
        ws.cell(row=next_row, column=4, value=ibkr_exit or None)
        ws.cell(row=next_row, column=5, value=tos_exit or None)
    left_align = Alignment(horizontal="left")
    for col in range(1, 6):
        ws.cell(row=1, column=col).alignment = left_align
    for row in range(2, ws.max_row + 1):
        for col in range(1, 6):
            ws.cell(row=row, column=col).alignment = left_align


def main() -> None:
    if not os.path.exists(SOURCE_FILE):
        print(f"Source Earnings file not found: {SOURCE_FILE}")
        return

    try:
        wb_source = load_workbook(SOURCE_FILE, data_only=True)
    except PermissionError:
        print("Please close Latest Earnings Document.")
        return

    try:
        ws_source = wb_source[SOURCE_SHEET]
    except KeyError:
        print(f"Worksheet '{SOURCE_SHEET}' not found in {SOURCE_FILE}.")
        return

    max_row = ws_source.max_row
    start_row = HEADER_ROW + 1

    # Collect by flag: "T" -> single-day list; 1,2,3,4 -> weekday lists
    t_rows: List[TradeRow] = []
    by_weekday: Dict[str, List[TradeRow]] = {
        "Monday": [], "Tuesday": [], "Wednesday": [], "Thursday": [], "Friday": []
    }

    for row in range(start_row, max_row + 1):
        flag_val = _parse_flag(ws_source[f"{COL_FLAG}{row}"].value)
        if flag_val is None:
            continue

        raw_ticker = ws_source[f"{COL_TICKER}{row}"].value
        raw_direction = ws_source[f"{COL_DIRECTION}{row}"].value
        raw_size = ws_source[f"{COL_SIZE}{row}"].value
        raw_ibkr_exit = ws_source[f"{COL_IBKR_EXIT}{row}"].value
        raw_tos_exit = ws_source[f"{COL_TOS_EXIT}{row}"].value

        ticker = _normalize_ticker(raw_ticker)
        direction = _normalize_direction(raw_direction)
        size = raw_size
        ibkr_exit = _cell_to_str(raw_ibkr_exit)
        tos_exit = _cell_to_str(raw_tos_exit)

        if not ticker:
            print(f"Row {row}: missing ticker; skipping.")
            continue
        if not direction:
            print(f"Row {row}: missing trade direction for ticker {ticker}; skipping.")
            continue
        if size is None or str(size).strip() == "":
            print(f"Row {row}: missing share size for ticker {ticker}; skipping.")
            continue

        row_data: TradeRow = (ticker, direction, size, ibkr_exit, tos_exit)
        if flag_val == "T":
            t_rows.append(row_data)
        else:
            day_name = _FLAG_TO_WEEKDAY[flag_val]
            by_weekday[day_name].append(row_data)

    # Any weekday with at least one row?
    has_weekday = any(by_weekday[d] for d in WEEKDAY_SHEETS)
    total_trades = len(t_rows) + sum(len(by_weekday[d]) for d in WEEKDAY_SHEETS)

    if total_trades == 0:
        print("No valid trades found (no rows with 'T' or 1/2/3/4/5 in column O had ticker, direction, and share size).")
        if os.path.exists(OUTPUT_FILE):
            try:
                wb_out = load_workbook(OUTPUT_FILE)
                if DAILY_SHEET in wb_out.sheetnames:
                    ws_out = wb_out[DAILY_SHEET]
                    if ws_out.max_row > 1:
                        ws_out.delete_rows(2, ws_out.max_row - 1)
                    wb_out.save(OUTPUT_FILE)
                    print("Cleared existing trades in Live_Trade_Info (Daily_Trades).")
            except PermissionError:
                print("Could not update Live_Trade_Info (file may be open). Please close it.")
                sys.exit(1)
        sys.exit(0)

    if has_weekday:
        print("No clear single day range, writing weekday trades.")
        # Multi-day: write only weekday sheets (ignore T rows)
        if not os.path.exists(OUTPUT_FILE):
            wb_output = Workbook()
            wb_output.remove(wb_output.active)
        else:
            wb_output = load_workbook(OUTPUT_FILE)
        for day_name in WEEKDAY_SHEETS:
            trades_for_day = by_weekday[day_name]
            if day_name not in wb_output.sheetnames:
                wb_output.create_sheet(day_name)
            ws = wb_output[day_name]
            _write_trades_to_sheet(ws, trades_for_day)
            if trades_for_day:
                print(f"Wrote {len(trades_for_day)} trade(s) to sheet '{day_name}'.")
        wb_output.save(OUTPUT_FILE)
        all_tickers = []
        for day_name in WEEKDAY_SHEETS:
            for t in by_weekday[day_name]:
                all_tickers.append(t[0])
        print(f"Tickers: {', '.join(all_tickers)}.")
    else:
        # Single-day: only T rows -> Daily_Trades
        print(f"Collected {len(t_rows)} trade(s) from Earnings (column O = T).")
        if os.path.exists(OUTPUT_FILE):
            wb_output = load_workbook(OUTPUT_FILE)
            ws_output = wb_output[DAILY_SHEET] if DAILY_SHEET in wb_output.sheetnames else wb_output.active
        else:
            wb_output = Workbook()
            ws_output = wb_output.active
        ws_output.title = DAILY_SHEET
        _write_trades_to_sheet(ws_output, t_rows)
        wb_output.save(OUTPUT_FILE)
        print(f"Wrote {len(t_rows)} trade(s) to '{OUTPUT_FILE}' (sheet '{DAILY_SHEET}').")
        print(f"Tickers: {', '.join(t[0] for t in t_rows)}.")


if __name__ == "__main__":
    main()
