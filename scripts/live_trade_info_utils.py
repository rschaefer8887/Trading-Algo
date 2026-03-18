"""
Shared helper for Live_Trade_Info.xlsx: resolve which sheet to use for trade data.

Reads the "Trade_Mode" sheet, cell C3. If the value is "Single Day", returns
"Daily_Trades"; otherwise returns the current weekday name (Monday, Tuesday, etc.)
so scripts can use Monday/Tuesday/Wednesday/Thursday sheets for multi-day mode.

If the file or sheet is missing, or the cell cannot be read, returns "Daily_Trades".
"""

import os
from datetime import datetime

TRADE_MODE_SHEET = "Trade_Mode"
TRADE_MODE_CELL = "C3"
MODE_SINGLE_DAY = "Single Day"
DEFAULT_SHEET = "Daily_Trades"


def get_trade_sheet_name(file_path: str) -> str:
    """
    Return the Live_Trade_Info sheet name to use for reading or writing trade data.

    - If Trade_Mode!C3 (stripped, case-insensitive) is "Single Day" -> "Daily_Trades".
    - Otherwise -> current weekday name (e.g. "Monday", "Tuesday") for multi-day mode.
    - If file missing, sheet missing, or read error -> "Daily_Trades".
    """
    if not file_path or not os.path.exists(file_path):
        return DEFAULT_SHEET
    openpyxl_value = None
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
        try:
            if TRADE_MODE_SHEET not in wb.sheetnames:
                openpyxl_value = None
            else:
                ws = wb[TRADE_MODE_SHEET]
                openpyxl_value = ws[TRADE_MODE_CELL].value
        finally:
            try:
                wb.close()
            except Exception:
                pass
    except Exception:
        openpyxl_value = None

    # Excel dropdown values sometimes come through as None via openpyxl.
    # If we didn't get a value, fall back to xlwings (evaluated by Excel).
    value = openpyxl_value
    if value is None:
        try:
            import xlwings as xw  # type: ignore

            app = xw.App(visible=False)
            wb = None
            try:
                wb = app.books.open(file_path)
                value = wb.sheets[TRADE_MODE_SHEET].range(TRADE_MODE_CELL).value
            finally:
                if wb is not None:
                    try:
                        wb.close()
                    except Exception:
                        pass
                try:
                    app.quit()
                except Exception:
                    pass
        except Exception:
            value = None

    if value is None:
        return DEFAULT_SHEET
    if str(value).strip().lower() == MODE_SINGLE_DAY.lower():
        return DEFAULT_SHEET
    return datetime.now().strftime("%A")
