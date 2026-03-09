"""
Send_Trades_Schwab — Send equity entry trades to Schwab based on Live_Trade_Info.xlsx.

Workflow:
- Reads Live_Trade_Info.xlsx (sheet "Prices"):
    A: Ticker (e.g. AAPL)
    B: Direction ("long" / "short")
    C: Share Size (integer)
    D: IBKR Exit (ignored here)
    E: ToS Exit (ignored here for now)
- For each valid row:
    - long  -> Schwab equity BUY market
    - short -> Schwab equity SELL SHORT market
- Prints a planned orders summary, then prompts:
    "Send Schwab live trades? (y/n)"
- If you confirm, sends orders via Schwab Trader API using schwab-py.

Prerequisites:
- schwab-py installed:
    python -m pip install --upgrade schwab-py
- Schwab_Auth.py configured and run at least once to complete OAuth:
    - Copy schwab_config.example.json -> schwab_config.json
    - Fill api_key, app_secret, callback_url, token_path, account_id
"""

import os
from typing import List, Tuple

from openpyxl import load_workbook

from Schwab_Auth import create_client  # same folder

try:
    from schwab.orders.equities import (
        equity_buy_market,
        equity_sell_short_market,
    )
except Exception as e:  # pragma: no cover - import-time failure
    equity_buy_market = None  # type: ignore[assignment]
    equity_sell_short_market = None  # type: ignore[assignment]
    SCHWAB_ORDERS_IMPORT_ERROR = e
else:
    SCHWAB_ORDERS_IMPORT_ERROR = None

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR = os.path.dirname(_SCRIPT_DIR)

LIVE_INFO_FILE = os.path.join(_BASE_DIR, "Live_Trade_Info.xlsx")
LIVE_INFO_SHEET = "Prices"


def _normalize_direction(direction_cell) -> str:
    if direction_cell is None:
        return ""
    return str(direction_cell).strip().lower()


def read_live_trades() -> List[Tuple[str, str, int]]:
    """
    Read Live_Trade_Info.xlsx and build a clean list of trades.

    Returns list of (ticker, direction, size).
    """
    if not os.path.exists(LIVE_INFO_FILE):
        raise FileNotFoundError(f"Live trade info file not found: {LIVE_INFO_FILE}")

    wb = load_workbook(LIVE_INFO_FILE, data_only=True)
    if LIVE_INFO_SHEET in wb.sheetnames:
        ws = wb[LIVE_INFO_SHEET]
    else:
        ws = wb.active

    trades: List[Tuple[str, str, int]] = []

    # Expect header in row 1: A=Ticker, B=Direction, C=Share Size
    for row in range(2, ws.max_row + 1):
        ticker_cell = ws.cell(row=row, column=1).value
        direction_cell = ws.cell(row=row, column=2).value
        size_cell = ws.cell(row=row, column=3).value

        if ticker_cell is None or str(ticker_cell).strip() == "":
            continue

        ticker = str(ticker_cell).strip().upper()
        direction_norm = _normalize_direction(direction_cell)

        if direction_norm not in ("long", "short"):
            print(f"Row {row}: invalid direction '{direction_cell}' for ticker {ticker}; skipping.")
            continue

        try:
            size = int(size_cell)
        except (TypeError, ValueError):
            print(f"Row {row}: invalid share size '{size_cell}' for ticker {ticker}; skipping.")
            continue

        if size <= 0:
            print(f"Row {row}: non-positive share size {size} for ticker {ticker}; skipping.")
            continue

        trades.append((ticker, direction_norm, size))

    return trades


def build_orders(trades: List[Tuple[str, str, int]]):
    """
    Build schwab-py equity order builders for each trade.

    For now:
      - long  -> equity_buy_market
      - short -> equity_sell_short_market
    """
    if SCHWAB_ORDERS_IMPORT_ERROR is not None or equity_buy_market is None:
        raise ImportError(
            "Could not import Schwab equity order templates from schwab-py.\n"
            "Install/update schwab-py with:\n"
            "    python -m pip install --upgrade schwab-py\n"
            f"Underlying import error: {SCHWAB_ORDERS_IMPORT_ERROR}"
        )

    orders = []
    for ticker, direction, size in trades:
        if direction == "long":
            ob = equity_buy_market(ticker, size)
        else:  # short
            ob = equity_sell_short_market(ticker, size)
        orders.append((ticker, direction, size, ob))
    return orders


def main() -> None:
    try:
        trades = read_live_trades()
    except FileNotFoundError as e:
        print(e)
        return

    if not trades:
        print("No valid trades found in Live_Trade_Info; nothing to send to Schwab.")
        return

    orders = build_orders(trades)

    print("\nPlanned Schwab entry trades:")
    for ticker, direction, size, _ in orders:
        side = "BUY" if direction == "long" else "SELL SHORT"
        print(f"  {side} {size} {ticker}")

    reply = input("\nSend Schwab live trades? (y/n): ").strip().lower()
    if reply not in ("y", "yes"):
        print("Exiting without sending Schwab trades.")
        return

    try:
        client, cfg = create_client()
    except Exception as e:
        print(f"Failed to create Schwab client: {e}")
        return

    account_id = cfg.get("account_id")
    if not account_id:
        print("account_id is missing from schwab_config.json; cannot place orders.")
        return

    print(f"\nPlacing orders to Schwab account {account_id} ...")
    for ticker, direction, size, ob in orders:
        try:
            order_spec = ob.build()
            resp = client.place_order(account_id, order_spec)
            print(f"Submitted {direction} {size} {ticker}, response: {resp.status_code if hasattr(resp, 'status_code') else resp}")
        except Exception as e:
            print(f"Error placing order for {ticker}: {e}")


if __name__ == "__main__":
    main()

