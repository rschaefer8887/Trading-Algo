"""
Write Opening Prices via Alpha Vantage

Reads tickers from the Latest Earnings workbook, fetches opening prices via
the Alpha Vantage API, writes to Opening_Prices_One_Time.xlsx and into the
Earnings workbook column T (same row as each ticker).

The Earnings file is opened and saved via xlwings (Excel) so that external
links and other Excel features are preserved. Close the workbook in Excel
before running.
"""

import os
import time
from datetime import datetime

from openpyxl import Workbook, load_workbook

try:
    import requests
except ImportError:
    requests = None

try:
    import xlwings as xw
except ImportError:
    xw = None

# ---------------------------------------------------------------------------
# Paths: Excel files live in the Trading Algo folder (parent of this script)
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR = os.path.dirname(_SCRIPT_DIR)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
source_file = os.path.join(_BASE_DIR, "! -- Latest Earnings Document.xlsx")
source_sheet = "Trades"
header_row = 3
ticker_column_letter = "A"

# Column in the Earnings workbook where opening price will be written (same row as ticker)
opening_price_column_letter = "T"

output_file = os.path.join(_BASE_DIR, "Opening_Prices_One_Time.xlsx")
output_sheet = "Prices"

# Alpha Vantage configuration
ALPHA_VANTAGE_API_KEY = "76J7TF0KG7G45VXM"
ALPHA_VANTAGE_URL = "https://www.alphavantage.co/query"

# Free tier: 5 requests per minute, 25 per day
MAX_TICKERS_PER_DAY = 25
SECONDS_BETWEEN_REQUESTS = 13  # ~4.6 requests/minute, under the 5/min limit


def fetch_open_price_alpha_vantage(ticker: str) -> float | None:
    """
    Fetch the most recent daily opening price for a single ticker from Alpha Vantage.

    Returns the price as a float rounded to 2 decimals, or None if unavailable.
    """
    if requests is None:
        print("The 'requests' library is not installed. Install it with: pip install requests")
        return None

    params = {
        "function": "TIME_SERIES_DAILY",
        "symbol": ticker,
        "outputsize": "compact",
        "apikey": ALPHA_VANTAGE_API_KEY,
    }

    try:
        resp = requests.get(ALPHA_VANTAGE_URL, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"HTTP error for {ticker}: {e}")
        return None

    # Handle rate limit / error messages from Alpha Vantage
    if "Note" in data:
        # Rate limit or other informational note; show it once and abort further calls.
        print(f"Alpha Vantage notice while fetching {ticker}: {data.get('Note')}")
        return None
    if "Error Message" in data:
        print(f"Alpha Vantage error for {ticker}: {data.get('Error Message')}")
        return None

    ts = data.get("Time Series (Daily)")
    if not isinstance(ts, dict) or not ts:
        print(f"No 'Time Series (Daily)' data for {ticker}. Raw keys: {list(data.keys())}")
        return None

    try:
        latest_date = sorted(ts.keys())[-1]
        open_str = ts[latest_date].get("1. open")
        if open_str is None:
            print(f"No '1. open' field for {ticker} on {latest_date}.")
            return None
        price = float(open_str)
        return round(price, 2)
    except Exception as e:
        print(f"Could not parse opening price for {ticker}: {e}")
        return None


def main() -> None:
    # -----------------------------------------------------------------------
    # Get the row range to process
    # -----------------------------------------------------------------------
    try:
        start_row = int(
            input(
                f"Enter starting row number to process (first ticker is row {header_row+1}): "
            )
        )
        end_row = int(input("Enter ending row number to process: "))
    except ValueError:
        print("Row numbers must be integers.")
        return

    if end_row < start_row:
        print("Ending row must be >= starting row.")
        return

    # -----------------------------------------------------------------------
    # Use xlwings to open Earnings in Excel, read tickers, write opening
    # prices, and save (Excel does the save so external links are preserved).
    # -----------------------------------------------------------------------
    if xw is None:
        print("xlwings is not installed. Install it with: pip install xlwings")
        return

    if not os.path.exists(source_file):
        print(f"Earnings file not found: {source_file}")
        return

    if requests is None:
        print("The 'requests' library is not installed. Install it with: pip install requests")
        return

    app = None
    tickers: list[str] = []
    price_map: dict[str, float | None] = {}

    try:
        app = xw.App(visible=False)
        wb_earnings = app.books.open(os.path.abspath(source_file))
        try:
            sheet = wb_earnings.sheets[source_sheet]
        except Exception:
            print(f"Worksheet '{source_sheet}' not found in {source_file}.")
            wb_earnings.close()
            return

        # Build list of (row, ticker_normalized) for rows that have a ticker
        row_tickers: list[tuple[int, str]] = []
        for row in range(start_row, end_row + 1):
            cell = sheet.range(f"{ticker_column_letter}{row}").value
            if cell is None:
                continue
            ticker = str(cell).strip().upper().replace(".", "-")
            row_tickers.append((row, ticker))

        if not row_tickers:
            print("No tickers found in the specified rows.")
            wb_earnings.close()
            return

        tickers = [t for _, t in row_tickers]
        unique_tickers = list(dict.fromkeys(tickers))  # preserve order, remove duplicates

        if len(unique_tickers) > MAX_TICKERS_PER_DAY:
            print(
                f"Warning: Alpha Vantage free tier allows {MAX_TICKERS_PER_DAY} requests per day.\n"
                f"Found {len(unique_tickers)} unique tickers. "
                f"Only the first {MAX_TICKERS_PER_DAY} will be requested; others will be left blank."
            )
            unique_tickers = unique_tickers[:MAX_TICKERS_PER_DAY]

        print(
            f"Found {len(tickers)} tickers ("
            f"{len(unique_tickers)} unique). Fetching opening prices from Alpha Vantage..."
        )
        print(
            "This may take several minutes due to free-tier rate limits "
            f"(sleeping {SECONDS_BETWEEN_REQUESTS} seconds between tickers)."
        )

        # Fetch opening prices for each unique ticker with rate limiting
        for idx, ticker in enumerate(unique_tickers):
            price = fetch_open_price_alpha_vantage(ticker)
            price_map[ticker] = price

            # Respect rate limit between calls, except after the last one
            if idx < len(unique_tickers) - 1:
                time.sleep(SECONDS_BETWEEN_REQUESTS)

        # For any ticker not fetched (e.g., beyond daily limit), leave as None
        for ticker in tickers:
            price_map.setdefault(ticker, None)

        # Write opening prices to column T (Excel saves the file → preserves external links)
        for row, ticker in row_tickers:
            val = price_map.get(ticker)
            sheet.range(f"{opening_price_column_letter}{row}").value = val

        wb_earnings.save()
        wb_earnings.close()
        print(
            f"Wrote opening prices to column {opening_price_column_letter} in "
            f"'{source_file}' and saved (via Excel)."
        )
    finally:
        if app is not None:
            app.quit()

    # -----------------------------------------------------------------------
    # Also write to Opening_Prices_One_Time.xlsx (same structure as original)
    # -----------------------------------------------------------------------
    if os.path.exists(output_file):
        wb_output = load_workbook(output_file)
        ws_output = wb_output.active
        ws_output.title = output_sheet
        ws_output.delete_rows(1, ws_output.max_row)
    else:
        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = output_sheet

    ws_output.append(["Ticker", "Date", "Opening Price"])
    today = datetime.today().strftime("%Y-%m-%d")

    for ticker in tickers:
        price = price_map.get(ticker)
        ws_output.append([ticker, today, price])

    wb_output.save(output_file)
    print(f"Successfully wrote {len(tickers)} tickers to '{output_file}'")


if __name__ == "__main__":
    main()

